import openpyxl
def getrow(xlpath, sheet):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.max_row

def getcolumn(xlpath, sheet):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.max_column

def readdata(xlpath, sheet, r, c):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.cell(r, c).value

def writedata(xlpath, sheet, r, c, data):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    sh.cell(r, c).value = data
    wb.save(xlpath)